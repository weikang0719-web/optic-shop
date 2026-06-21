from flask import send_file
from io import BytesIO
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas
from flask import Flask, request, session, redirect
from datetime import datetime
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Spacer,
    Paragraph
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import calendar
import sqlite3
import os
import psycopg2

app = Flask(__name__)
app.secret_key = "optic_shop_secret_key"

DATABASE_URL = os.environ.get("DATABASE_URL")

def get_conn():
    return psycopg2.connect(DATABASE_URL)

def init_db():
    conn = get_conn()
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS sales (
            id SERIAL PRIMARY KEY,
            customer TEXT,
            amount REAL,
            staff TEXT,
            date TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        ALTER TABLE sales
        ADD COLUMN IF NOT EXISTS company_code TEXT
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id SERIAL PRIMARY KEY,
            category TEXT,
            amount REAL,
            note TEXT,
            date TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        ALTER TABLE expenses
        ADD COLUMN IF NOT EXISTS company_code TEXT
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS salaries (
            id SERIAL PRIMARY KEY,
            staff TEXT,
            amount REAL,
            month TEXT,
            date TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        ALTER TABLE salaries
        ADD COLUMN IF NOT EXISTS company_code TEXT
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE,
            password TEXT,
            role TEXT DEFAULT 'staff',

            can_add_sales BOOLEAN DEFAULT FALSE,
            can_edit_sales BOOLEAN DEFAULT FALSE,
            can_delete_sales BOOLEAN DEFAULT FALSE,

            can_add_expenses BOOLEAN DEFAULT FALSE,
            can_edit_expenses BOOLEAN DEFAULT FALSE,
            can_delete_expenses BOOLEAN DEFAULT FALSE,

            can_add_salary BOOLEAN DEFAULT FALSE,
            can_edit_salary BOOLEAN DEFAULT FALSE,
            can_delete_salary BOOLEAN DEFAULT FALSE,

            can_view_reports BOOLEAN DEFAULT FALSE,
            can_export BOOLEAN DEFAULT FALSE,
            can_backup BOOLEAN DEFAULT FALSE,
            can_restore BOOLEAN DEFAULT FALSE,

        is_active BOOLEAN DEFAULT TRUE
        )
    """)

    permission_columns = [
    "can_add_sales",
    "can_edit_sales",
    "can_delete_sales",
    "can_add_expenses",
    "can_edit_expenses",
    "can_delete_expenses",
    "can_add_salary",
    "can_edit_salary",
    "can_delete_salary",
    "can_view_reports",
    "can_export",
    "can_backup",
    "can_restore",
    "is_active"
    ]

    for col in permission_columns:
        c.execute(f"""
        ALTER TABLE users
        ADD COLUMN IF NOT EXISTS {col} BOOLEAN DEFAULT TRUE
    """)

    c.execute("""
    INSERT INTO users (
        username, password, role,
        can_add_sales, can_edit_sales, can_delete_sales,
        can_add_expenses, can_edit_expenses, can_delete_expenses,
        can_add_salary, can_edit_salary, can_delete_salary,
        can_view_reports, can_export, can_backup, can_restore,
        is_active
    )
    VALUES (
        'admin', 'admin123', 'admin',
        TRUE, TRUE, TRUE,
        TRUE, TRUE, TRUE,
        TRUE, TRUE, TRUE,
        TRUE, TRUE, TRUE, TRUE,
        TRUE
    )
    ON CONFLICT (username) DO UPDATE SET
        password='admin123',
        role='admin',
        can_add_sales=TRUE,
        can_edit_sales=TRUE,
        can_delete_sales=TRUE,
        can_add_expenses=TRUE,
        can_edit_expenses=TRUE,
        can_delete_expenses=TRUE,
        can_add_salary=TRUE,
        can_edit_salary=TRUE,
        can_delete_salary=TRUE,
        can_view_reports=TRUE,
        can_export=TRUE,
        can_backup=TRUE,
        can_restore=TRUE,
        is_active=TRUE
""")

    c.execute("""
    CREATE TABLE IF NOT EXISTS companies (
        id SERIAL PRIMARY KEY,
        company_code TEXT UNIQUE,
        company_name TEXT,
        address TEXT,
        phone TEXT,
        is_active BOOLEAN DEFAULT TRUE,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
    """)

    c.execute("""
    ALTER TABLE users
    ADD COLUMN IF NOT EXISTS company_code TEXT
    """)

    c.execute("""
    ALTER TABLE sales
    ADD COLUMN IF NOT EXISTS company_code TEXT
    """)

    c.execute("""
    ALTER TABLE sales
    ADD COLUMN IF NOT EXISTS receipt_no TEXT
    """)

    c.execute("""
    ALTER TABLE sales
    ADD COLUMN IF NOT EXISTS reference_no TEXT
    """)

    c.execute("""
    ALTER TABLE sales
    ADD COLUMN IF NOT EXISTS remarks TEXT
    """)

    c.execute("""
    ALTER TABLE sales
    ADD COLUMN IF NOT EXISTS payment_method TEXT
    """)

    c.execute("""
    ALTER TABLE expenses
    ADD COLUMN IF NOT EXISTS company_code TEXT
    """)

    c.execute("""
    ALTER TABLE salaries
    ADD COLUMN IF NOT EXISTS company_code TEXT
    """)

    c.execute("""
    ALTER TABLE companies
    ADD COLUMN IF NOT EXISTS email TEXT
    """)

    c.execute("""
    ALTER TABLE companies
    ADD COLUMN IF NOT EXISTS receipt_footer TEXT
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS support_tickets (
        id SERIAL PRIMARY KEY,
        company_code TEXT,
        username TEXT,
        subject TEXT,
        description TEXT,
        status TEXT DEFAULT 'OPEN',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS stock (
        id SERIAL PRIMARY KEY,
        company_code TEXT,
        item_code TEXT,
        item_name TEXT,
        cost REAL DEFAULT 0,
        commission REAL DEFAULT 0,
        minimum_selling_price REAL DEFAULT 0,
        supplier TEXT,
        qty INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS suppliers (
        id SERIAL PRIMARY KEY,
        company_code TEXT,
        supplier_code TEXT,
        supplier_name TEXT,
        phone TEXT,
        address TEXT,
        account_no TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS stock_movements (
        id SERIAL PRIMARY KEY,
        company_code TEXT,
        movement_date TEXT,
        movement_type TEXT,
        reference_no TEXT,
        supplier_id INTEGER,
        item_id INTEGER,
        qty INTEGER DEFAULT 0,
        note TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
    """)

    c.execute("""
    ALTER TABLE stock_movements
    ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'ACTIVE'
    """)

    c.execute("""
    ALTER TABLE companies
    ADD COLUMN IF NOT EXISTS low_stock_alert_enabled BOOLEAN DEFAULT FALSE
    """)

    c.execute("""
    ALTER TABLE companies
    ADD COLUMN IF NOT EXISTS low_stock_threshold INTEGER DEFAULT 3
    """)

    c.execute("""
    ALTER TABLE companies
    ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'ACTIVE'
    """)

    c.execute("""
    ALTER TABLE companies
    ADD COLUMN IF NOT EXISTS expiry_date DATE
    """)

    c.execute("""
    UPDATE companies
    SET status='ACTIVE'
    WHERE status IS NULL
    """)

    c.execute("""
    ALTER TABLE stock
    ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE
    """)

    c.execute("""
    ALTER TABLE stock
    ADD COLUMN IF NOT EXISTS selling_price NUMERIC(12,2) DEFAULT 0
    """)

    c.execute("""
    UPDATE stock
    SET is_active=TRUE
    WHERE is_active IS NULL
    """)

    c.execute("""
    ALTER TABLE suppliers
    ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE
    """)

    c.execute("""
    UPDATE suppliers
    SET is_active=TRUE
    WHERE is_active IS NULL
    """)

    c.execute("""
    ALTER TABLE suppliers
    ADD COLUMN IF NOT EXISTS tel_no TEXT
    """)

    c.execute("""
    ALTER TABLE suppliers
    ADD COLUMN IF NOT EXISTS address TEXT
    """)

    c.execute("""
    ALTER TABLE suppliers
    ADD COLUMN IF NOT EXISTS account_no TEXT
    """)

    c.execute("""
    ALTER TABLE suppliers
    ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS stock_adjustments (
        id SERIAL PRIMARY KEY,
        company_code TEXT,
        adjustment_date DATE,
        item_id INTEGER,
        qty_change INTEGER,
        reason TEXT,
        created_by TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    conn.commit()
    conn.close()

init_db()

@app.route("/register-owner", methods=["GET", "POST"])
def register_owner():

    if request.method == "POST":
        company_code = request.form["company_code"]
        username = request.form["username"]
        password = request.form["password"]

        conn = get_conn()
        c = conn.cursor()

        c.execute("""
            SELECT company_code 
            FROM companies 
            WHERE company_code=%s AND is_active=TRUE
        """, (company_code,))

        company = c.fetchone()

        if not company:
            conn.close()
            return "Invalid Company Code"

        c.execute("""
            INSERT INTO users (username, password, role, company_code, is_active)
            VALUES (%s, %s, 'owner', %s, TRUE)
        """, (username, password, company_code))

        conn.commit()
        conn.close()

        return redirect("/login")

    return """
    <h1>Company Profile</h1>
    Coming Soon
    """

    return """
    <h1>Register Owner Account</h1>

    <form method="POST">
        Company Code:<br>
        <input name="company_code" required><br><br>

        Username:<br>
        <input name="username" required><br><br>

        Password:<br>
        <input type="password" name="password" required><br><br>

        <button type="submit">Create Owner Account</button>
    </form>

    <br>
    <a href="/login">Login</a>
    """

@app.route("/company-profile", methods=["GET", "POST"])
def company_profile():

    if not session.get("logged_in"):
        return redirect("/login")

    company_code = session["company_code"]

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":
        company_name = request.form["company_name"]
        address = request.form["address"]
        phone = request.form["phone"]
        email = request.form["email"]
        receipt_footer = request.form["receipt_footer"]
        low_stock_alert_enabled = "low_stock_alert_enabled" in request.form
        low_stock_threshold = int(request.form["low_stock_threshold"] or 3)

        c.execute("""
            UPDATE companies
            SET company_name=%s,
                address=%s,
                phone=%s,
                email=%s,
                receipt_footer=%s,
                low_stock_alert_enabled=%s,
                low_stock_threshold=%s
            WHERE company_code=%s
        """, (
            company_name,
            address,
            phone,
            email,
            receipt_footer,
            low_stock_alert_enabled,
            low_stock_threshold,
            session["company_code"]
        ))

        conn.commit()

    c.execute("""
        SELECT company_code, company_name, address, phone, email, receipt_footer,
        low_stock_threshold, low_stock_alert_enabled
        FROM companies
        WHERE company_code=%s
    """, (company_code,))

    company = c.fetchone()
    conn.close()

    checked_alert = "checked" if company[7] else ""

    return f"""
    <h1>Company Profile</h1>

    <form method="POST">

        Company Code<br>
        <input type="text" value="{company[0]}" readonly><br><br>

        Company Name<br>
        <input type="text" name="company_name" value="{company[1] or ''}" required><br><br>

        Address<br>
        <textarea name="address" rows="4" cols="50">{company[2] or ''}</textarea><br><br>

        Phone<br>
        <input type="text" name="phone" value="{company[3] or ''}"><br><br>

        Email<br>
        <input type="email" name="email" value="{company[4] or ''}"><br><br>

        Receipt Footer<br>
        <textarea name="receipt_footer" rows="3" cols="50">{company[5] or ''}</textarea><br><br>

        Low Stock Alert:<br>
        <label>
            <input type="checkbox" name="low_stock_alert_enabled" {checked_alert}>
            Enable Low Stock Alert
        </label><br><br>

        Low Stock Threshold:<br>
        <input type="number" name="low_stock_threshold" value="{company[6] or 3}" min="0"><br><br>

        <button type="submit">Save</button>

    </form>

    <br>
    <a href="/">Back to Dashboard</a>
    """

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_conn()
        c = conn.cursor()

        c.execute("""
            SELECT
                username, role,
                can_add_sales, can_edit_sales, can_delete_sales,
                can_add_expenses, can_edit_expenses, can_delete_expenses,
                can_add_salary, can_edit_salary, can_delete_salary,
                can_view_reports, can_export, can_backup, can_restore, company_code
            FROM users
            WHERE username=%s AND password=%s AND is_active=TRUE
        """, (username, password))

        user = c.fetchone()
        conn.close()

        if user:
            session["logged_in"] = True
            session["username"] = user[0]
            session["role"] = user[1]
            session["can_add_sales"] = user[2]
            session["can_edit_sales"] = user[3]
            session["can_delete_sales"] = user[4]
            session["can_add_expenses"] = user[5]
            session["can_edit_expenses"] = user[6]
            session["can_delete_expenses"] = user[7]
            session["can_add_salary"] = user[8]
            session["can_edit_salary"] = user[9]
            session["can_delete_salary"] = user[10]
            session["can_view_reports"] = user[11]
            session["can_export"] = user[12]
            session["can_backup"] = user[13]
            session["can_restore"] = user[14]
            session["company_code"] = user[15]

            if session["role"] == "admin":
                return redirect("/admin")

            return redirect("/")
        else:
            return """
            <h1>Login Failed</h1>
            <p>Wrong username or password.</p>
            <a href="/login">Try Again</a>
            """

    return """
    <!DOCTYPE html>
    <html>
<head>
<title>Login</title>
<style>
body{
    font-family: Arial;
    background:#f4f6f9;
    display:flex;
    justify-content:center;
    align-items:center;
    height:100vh;
}

.login-box{
    background:white;
    padding:30px;
    border-radius:10px;
    width:350px;
    box-shadow:0 0 15px rgba(0,0,0,0.1);
}

h2{
    text-align:center;
    margin-bottom:20px;
}

input{
    width:100%;
    padding:10px;
    margin-top:5px;
    margin-bottom:15px;
    border:1px solid #ccc;
    border-radius:5px;
}

button{
    width:100%;
    padding:10px;
    background:#007bff;
    color:white;
    border:none;
    border-radius:5px;
    cursor:pointer;
}

button:hover{
    background:#0056b3;
}
</style>
</head>

<body>

<div class="login-box">
    <h2>Optic Shop Login</h2>

    <form method="POST">

        <label>Username</label>
        <input type="text" name="username" required>

        <label>Password</label>
        <input type="password" name="password" required>

        <button type="submit">Login</button>

    </form>
</div>

</body>
</html>
"""

def has_permission(permission):
    return session.get(permission, False) or session.get("role") == "admin"

@app.route("/")
def home():
    if not session.get("logged_in"):
        return redirect("/login")
    
    if session.get("role") == "admin":
        return redirect("/admin")

    conn = get_conn()
    c = conn.cursor()

    from datetime import datetime

    current_month = datetime.now().strftime("%Y-%m")

    # Monthly Sales
    if session.get("role") == "admin":
        c.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM sales
        WHERE date LIKE %s
        """, (current_month + "%",))
    else:
        c.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM sales
        WHERE date LIKE %s AND company_code=%s
        """, (current_month + "%", session["company_code"]))

    total_sales = c.fetchone()[0]

    # Monthly Expenses
    if session.get("role") == "admin":
        c.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM expenses
        WHERE date LIKE %s
        """, (current_month + "%",))
    else:
        c.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM expenses
        WHERE date LIKE %s AND company_code=%s
        """, (current_month + "%", session["company_code"]))

    normal_expenses = c.fetchone()[0]

    # Monthly Salaries
    if session.get("role") == "admin":
        c.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM salaries
        WHERE date LIKE %s
        """, (current_month + "%",))
    else:
        c.execute("""
        SELECT COALESCE(SUM(amount),0)
        FROM salaries
        WHERE date LIKE %s AND company_code=%s
        """, (current_month + "%", session["company_code"]))

    total_salaries = c.fetchone()[0]

    total_expenses = normal_expenses + total_salaries
    profit = total_sales - total_expenses
    profit_color = "green"
    if profit < 0:
        profit_color = "red"

    c.execute("""
        SELECT expiry_date
        FROM companies
        WHERE company_code=%s
    """, (session["company_code"],))

    company_row = c.fetchone()
    expiry_date = company_row[0] if company_row else None
    
    conn.close()

    menu_html = ""

    if session.get("can_add_sales"):
        menu_html += '<a href="/sales"><button>Add Sales</button></a>'

    if session.get("can_add_expenses"):
        menu_html += '<a href="/expenses"><button>Add Expense</button></a>'

    if session.get("can_add_salary"):
        menu_html += '<a href="/salary"><button>Add Staff Salary</button></a>'

    if session.get("can_view_reports"):
        menu_html += '<a href="/reports"><button>Reports</button></a>'

    if session.get("role") in ["admin", "owner"]:
        menu_html += '<a href="/permissions"><button>Permissions</button></a>'
        menu_html += '<a href="/company-profile"><button>Company Profile</button></a>'
        menu_html += '<a href="/stock"><button>Stock</button></a>'
        menu_html += '<a href="/suppliers"><button>Suppliers</button></a>'
        menu_html += '<a href="/stock-movement"><button>Stock Purchase In / Out</button></a>'
        menu_html += '<a href="/stock-adjustment"><button>Stock Adjustment</button></a>'
        menu_html += '<a href="/stock-balance"><button>Check Stock Balance</button></a>'

    menu_html += '<a href="/support"><button>Report Problem</button></a>'
    
    menu_html += '<a href="/logout"><button>Logout</button></a>'

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Optic Shop System</title>
        <style>
body {{
    font-family: Arial;
    background: #f4f6f8;
    padding: 30px;
}}

.header {{
    background: #111827;
    color: white;
    padding: 20px;
    border-radius: 10px;
    margin-bottom: 25px;
}}

.cards {{
    display: flex;
    gap: 20px;
    flex-wrap: wrap;
    margin-bottom: 30px;
}}

.card {{
    background: white;
    padding: 25px;
    border-radius: 10px;
    width: 220px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
}}

.card h3 {{
    margin: 0;
    color: #555;
}}

.amount {{
    font-size: 28px;
    font-weight: bold;
    margin-top: 10px;
}}

.menu {{
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
}}

button {{
    background: #111827;
    color: white;
    border: none;
    padding: 12px 18px;
    border-radius: 8px;
    cursor: pointer;
}}

button:hover {{
    background: #2563eb;
}}
</style>
    </head>

    <body>
        <div class="header">
            <h1>OPTIC SHOP MANAGEMENT SYSTEM</h1>
            <p>Business Dashboard</p>

            <p style="
                color:#ffcc00;
                font-size:18px;
                font-weight:bold;
            ">
                Expiry Date: {expiry_date}
            </p>

            <div id="datetime" style="
                color:white;
                font-size:18px;
                font-weight:bold;
                margin-top:10px;
        "></div>
        </div>

        <div class="cards">
            <div class="card">
                <h3>This Month Sales</h3>
                <div class="amount">RM {total_sales:,.2f}</div>
            </div>

            <div class="card">
                <h3>This Month Expenses</h3>
                <div class="amount">RM {total_expenses:,.2f}</div>
            </div>

            <div class="card">
                <h3>This Month P&L</h3>
                <div class="amount" style="color:{profit_color}">RM {profit:,.2f}
            </div>
        </div>

        <div class="menu">
            {menu_html}
        </div>

        <script>
            function updateDateTime() {{
            const now = new Date();

            const date = now.toLocaleDateString('en-GB', {{
                weekday: 'long',
                year: 'numeric',
                month: 'long',
                day: 'numeric'
            }});

    const time = now.toLocaleTimeString('en-GB');

    document.getElementById("datetime").innerHTML =
        date + "<br>" + time;
}}

updateDateTime();
setInterval(updateDateTime, 1000);
</script>

</body>
</html>
"""

@app.route("/sales", methods=["GET", "POST"])
def sales():
    if not session.get("logged_in"):
        return redirect("/login")
    
    if not has_permission("can_add_sales"):
        return "Access Denied"

    if request.method == "POST":
        sale_date = request.form["sale_date"]
        customer = request.form["customer"]
        reference_no = request.form["reference_no"]
        amount = float(request.form["amount"])
        staff = request.form["staff"]
        remarks = request.form["remarks"]
        payment_method = request.form["payment_method"]

        conn = get_conn()
        c = conn.cursor()

        c.execute("""
            INSERT INTO sales (
                date,
                customer,
                reference_no,
                remarks,
                payment_method,
                amount,
                staff,
                company_code
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            sale_date,
            customer,
            reference_no,
            remarks,
            payment_method,
            amount,
            staff,
            session["company_code"]
        ))

        sale_id = c.fetchone()[0]

        c.execute("""
            SELECT COUNT(*)
            FROM sales
            WHERE company_code=%s
        """, (session["company_code"],))

        receipt_count = c.fetchone()[0]

        receipt_no = f"REC-{receipt_count:06d}"

        c.execute("""
            UPDATE sales
            SET receipt_no=%s
            WHERE id=%s
        """, (receipt_no, sale_id))

        conn.commit()
        conn.close()

        return redirect(f"/receipt/{sale_id}")

    return f"""
    <h1>Add Sales</h1>

    <form method="POST">

        <label>Date:</label><br>
        <input type="date" name="sale_date" value="{datetime.now().strftime('%Y-%m-%d')}" required><br><br>

        <label>Customer:</label><br>
        <input type="text" name="customer" required><br><br>

        <label>Reference No:</label><br>
        <input type="text" name="reference_no"><br><br>

        <label>Remark:</label><br>
        <textarea name="remarks" rows="3"></textarea><br><br>

        <label>Payment Method:</label><br>
        <select name="payment_method" required>
            <option value="Cash">Cash</option>
            <option value="Debit Card">Debit Card</option>
            <option value="Credit Card">Credit Card</option>
            <option value="DuitNow">DuitNow</option>
            <option value="Atome">Atome</option>
        </select><br><br>

        <label>Amount (RM):</label><br>
        <input type="number" step="0.01" name="amount" required><br><br>

        <label>Staff:</label><br>
        <input type="text" name="staff" required><br><br>

        <button type="submit">Save</button>

        <button type="button"
        onclick="window.location.href='/'">
           Cancel
        </button>

    </form>

    <br>
    <a href="/">Back to Dashboard</a>
    """


@app.route("/sales-list")
def sales_list():
    if not session.get("logged_in"):
        return redirect("/login")

    today = datetime.now()
    first_day = today.replace(day=1).strftime("%Y-%m-%d")
    last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1]).strftime("%Y-%m-%d")

    from_date = request.args.get("from_date", first_day)
    to_date = request.args.get("to_date", last_day)

    conn = get_conn()
    c = conn.cursor()

    if session.get("role") == "admin":
        c.execute("""
            SELECT id, customer, amount, staff, date, receipt_no, reference_no, payment_method
            FROM sales
            WHERE date BETWEEN %s AND %s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date))
    else:
        c.execute("""
            SELECT id, customer, amount, staff, date, receipt_no, reference_no, payment_method
            FROM sales
            WHERE date BETWEEN %s AND %s
            AND company_code=%s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date, session["company_code"]))

    records = c.fetchall()
    conn.close()

    rows = ""
    for r in records:
        rows += f"""
        <tr>
            <td>{str(r[4])[:10]}</td>
            <td>{r[5] or ('REC-' + str(r[0]).zfill(6))}</td>
            <td>{r[1]}</td>
            <td>{r[6] or '-'}</td>
            <td>{r[7] or '-'}</td>
            <td style="text-align:left; width:35px;">RM</td>
            <td style="text-align:right; width:110px;">{float(r[2]):,.2f}</td>
            <td>{r[3]}</td>
            <td>
                <a href="/edit-sale/{r[0]}">Edit</a> |
                <a href="/delete-sale/{r[0]}" onclick="return confirm('Delete this sale?')">Delete</a>
                <a href="/receipt/{r[0]}">Print</a>
            </td>
        </tr>
        """

    return f"""
    <h1>Sales List</h1>

    <form method="GET">
        <label>From Date:</label>
        <input type="date" name="from_date" value="{from_date}" required>

        <label>To Date:</label>
        <input type="date" name="to_date" value="{to_date}" required>

        <button type="submit">Search</button>
    </form>

    <br>

    <table border="1" cellpadding="10">
        <tr>
            <th>Date</th>
            <th>Receipt No</th>
            <th>Customer</th>
            <th>Ref No</th>
            <th>Payment</th>
            <th colspan="2">Amount</th>
            <th>Staff</th>
            <th>Action</th>
        </tr>
        {rows}
    </table>

    <br>
    <a href="/">Back to Dashboard</a>
    """

@app.route("/receipt/<int:sale_id>")
def receipt(sale_id):

    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    c.execute("""
        SELECT company_name, address, phone, email, receipt_footer
        FROM companies
        WHERE company_code=%s
    """, (session["company_code"],))
    company = c.fetchone()

    c.execute("""
        SELECT id, date, customer, amount, staff, receipt_no, reference_no, remarks, payment_method
        FROM sales
        WHERE id=%s AND company_code=%s
    """, (sale_id, session["company_code"]))
    sale = c.fetchone()

    conn.close()

    if not sale:
        return "Receipt not found"

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Receipt #{sale[0]}</title>
        <style>
            body {{
                font-family: Arial;
                width: 300px;
                margin: 20px auto;
                font-size: 14px;
            }}
            .center {{
                text-align: center;
            }}
            .line {{
                border-top: 1px dashed #000;
                margin: 10px 0;
            }}
            .total {{
                font-size: 20px;
                font-weight: bold;
            }}
            @media print {{
                button, a {{
                    display: none;
                }}
            }}
        </style>
    </head>
    <body>

        <div class="center">
            <h2>{company[0] or ''}</h2>
            <p>{company[1] or ''}</p>
            <p>Tel: {company[2] or ''}</p>
            <p>Email: {company[3] or ''}</p>
        </div>

        <div class="line"></div>

        <p><b>Receipt No:</b> {sale[5] or ('REC-' + str(sale[0]).zfill(6))}</p>
        <p><b>Date:</b> {str(sale[1])[:10]}</p>
        <p><b>Ref No:</b> {sale[6] or '-'}</p>
        <p><b>Customer:</b> {sale[2]}</p>
        <p><b>Remark:</b> {sale[7] or '-'}</p>
        <p><b>Staff:</b> {sale[4]}</p>

        <div class="line"></div>

        <p class="total"><b>Amount:</b> RM {float(sale[3]):,.2f}</p>

        <div class="line"></div>

        <p><b>Payment Method:</b> {sale[8] or '-'}</p>

        <div class="center">
            <p>{company[4] or 'Thank you for your support.'}</p>
        </div>

        <br>

        <button onclick="window.print()">Print Receipt</button>
        <br><br>
        <a href="/sales-list">Back</a>

    </body>
    </html>
    """

@app.route("/edit-sale/<int:sale_id>", methods=["GET", "POST"])
def edit_sale(sale_id):
    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":
        sale_date = request.form["sale_date"]
        customer = request.form["customer"]
        amount = float(request.form["amount"].replace(",", ""))
        staff = request.form["staff"]

        c.execute("""
            UPDATE sales
            SET date=%s, customer=%s, amount=%s, staff=%s
            WHERE id=%s
        """, (sale_date, customer, amount, staff, sale_id))

        conn.commit()
        conn.close()
        return redirect("/sales-list")

    c.execute("SELECT id, customer, amount, staff, date FROM sales WHERE id=%s", (sale_id,))
    sale = c.fetchone()
    conn.close()

    return f"""
    <h1>Edit Sale</h1>

    <form method="POST">
        <label>Date:</label><br>
        <input type="date" name="sale_date" value="{str(sale[4])[:10]}" required><br><br>

        <label>Customer Name:</label><br>
        <input type="text" name="customer" value="{sale[1]}" required><br><br>

        <label>Sales Amount (RM):</label><br>
        <input type="text" name="amount" value="{float(sale[2]):,.2f}" required><br><br>

        <label>Staff Name:</label><br>
        <input type="text" name="staff" value="{sale[3]}" required><br><br>

        <button type="submit">Update Sale</button>
    </form>

    <br>
    <a href="/sales-list">Back to Sales List</a>
    """

@app.route("/salary", methods=["GET", "POST"])
def salary():
    if not session.get("logged_in"):
        return redirect("/login")
    
    if not has_permission("can_add_salary"):
        return "Access Denied"

    if request.method == "POST":
        salary_date = request.form["salary_date"]
        staff = request.form["staff"]
        amount = float(request.form["amount"].replace(",", ""))
        month = request.form["month"]

        conn = get_conn()
        c = conn.cursor()

        c.execute("""
        INSERT INTO salaries (
            date,
            staff,
            amount,
            month,
            company_code
        )
        VALUES (%s, %s, %s, %s, %s)
    """, (salary_date, staff, amount, month, session["company_code"]
    ))

        conn.commit()
        conn.close()

        return redirect("/salary-list")

    return f"""
    <h1>Add Staff Salary</h1>

    <form method="POST">

        <label>Date:</label><br>
        <input type="date" name="salary_date" value="{datetime.now().strftime('%Y-%m-%d')}" required><br><br>

        <label>Staff Name:</label><br>
        <input type="text" name="staff" required><br><br>

        <label>Salary Amount (RM):</label><br>
        <input type="text" name="amount" required><br><br>

        <label>Month:</label><br>
        <input type="month"
               name="month"
               value="{datetime.now().strftime('%Y-%m')}"
               required><br><br>

        <button type="submit">Save</button>

<button type="button"
        onclick="window.location.href='/'">
    Cancel
</button>

    </form>

    <br>
    <a href="/">Back to Dashboard</a>
    """

@app.route("/salary-list")
def salary_list():
    if not session.get("logged_in"):
        return redirect("/login")

    today = datetime.now()
    first_day = today.replace(day=1).strftime("%Y-%m-%d")
    last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1]).strftime("%Y-%m-%d")

    from_date = request.args.get("from_date", first_day)
    to_date = request.args.get("to_date", last_day)

    conn = get_conn()
    c = conn.cursor()

    if session.get("role") == "admin":
        c.execute("""
            SELECT id, staff, amount, month, date
            FROM salaries
            WHERE date BETWEEN %s AND %s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date))
    else:
        c.execute("""
            SELECT id, staff, amount, month, date
            FROM salaries
            WHERE date BETWEEN %s AND %s
            AND company_code=%s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date, session["company_code"]))

    records = c.fetchall()
    conn.close()

    rows = ""
    for r in records:
        rows += f"""
        <tr>
            <td>{str(r[4])[:10]}</td>
            <td>{r[1]}</td>
            <td>RM {float(r[2]):,.2f}</td>
            <td>{r[3]}</td>
            <td>
                <a href="/edit-salary/{r[0]}">Edit</a> |
                <a href="/delete-salary/{r[0]}" onclick="return confirm('Delete this salary?')">Delete</a>
            </td>
        </tr>
        """

    return f"""
    <h1>Salary List</h1>

    <form method="GET">
        <label>From Date:</label>
        <input type="date" name="from_date" value="{from_date}" required>

        <label>To Date:</label>
        <input type="date" name="to_date" value="{to_date}" required>

        <button type="submit">Search</button>
    </form>

    <br>

    <table border="1" cellpadding="10">
        <tr>
            <th>Date</th>
            <th>Staff</th>
            <th>Amount</th>
            <th>Month</th>
            <th>Action</th>
        </tr>
        {rows}
    </table>

    <br>
    <a href="/">Back to Dashboard</a>
    """

@app.route("/edit-salary/<int:salary_id>", methods=["GET", "POST"])
def edit_salary(salary_id):
    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":
        salary_date = request.form["salary_date"]
        staff = request.form["staff"]
        amount = float(request.form["amount"].replace(",", ""))
        month = request.form["month"]

        c.execute("""
            UPDATE salaries
            SET date=%s, staff=%s, amount=%s, month=%s
            WHERE id=%s
        """, (salary_date, staff, amount, month, salary_id))

        conn.commit()
        conn.close()
        return redirect("/salary-list")

    c.execute("SELECT id, staff, amount, month, date FROM salaries WHERE id=%s", (salary_id,))
    salary = c.fetchone()
    conn.close()

    return f"""
    <h1>Edit Staff Salary</h1>

    <form method="POST">
        <label>Date:</label><br>
        <input type="date" name="salary_date" value="{str(salary[4])[:10]}" required><br><br>

        <label>Staff Name:</label><br>
        <input type="text" name="staff" value="{salary[1]}" required><br><br>

        <label>Salary Amount (RM):</label><br>
        <input type="text" name="amount" value="{float(salary[2]):,.2f}" required><br><br>

        <label>Month:</label><br>
        <input type="month" name="month" value="{salary[3]}" required><br><br>

        <button type="submit">Update Salary</button>
    </form>

    <br>
    <a href="/salary-list">Back to Salary List</a>
    """

@app.route("/delete-salary/<int:salary_id>")
def delete_salary(salary_id):
    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM salaries WHERE id=%s", (salary_id,))
    conn.commit()
    conn.close()

    return redirect("/salary-list")

@app.route("/expenses", methods=["GET", "POST"])
def expense():
    if not session.get("logged_in"):
        return redirect("/login")

    if not has_permission("can_add_expenses"):
        return "Access Denied"

    if request.method == "POST":
        expense_date = request.form["expense_date"]
        category = request.form["category"]
        amount = float(request.form["amount"])
        note = request.form["note"]

        conn = get_conn()
        c = conn.cursor()

        c.execute("""
            INSERT INTO expenses (
                date,
                category,
                amount,
                note,
                company_code
            )
            VALUES (%s, %s, %s, %s, %s)
        """, (
            expense_date,
            category,
            amount,
            note,
            session["company_code"]
        ))

        conn.commit()
        conn.close()

        return redirect("/expenses-list")

    return f"""
    <h1>Add Expense</h1>

    <form method="POST">
        <label>Date:</label><br>
        <input type="date" name="expense_date" value="{datetime.now().strftime('%Y-%m-%d')}" required><br><br>

        <label>Category:</label><br>
        <input type="text" name="category" required><br><br>

        <label>Amount:</label><br>
        <input type="number" step="0.01" name="amount" required><br><br>

        <label>Note:</label><br>
        <input type="text" name="note"><br><br>

        <button type="submit">Save Expense</button>
    </form>

    <br>
    <a href="/">Back Dashboard</a>
    """


@app.route("/expenses-list")
def expenses_list():
    if not session.get("logged_in"):
        return redirect("/login")

    today = datetime.now()
    first_day = today.replace(day=1).strftime("%Y-%m-%d")
    last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1]).strftime("%Y-%m-%d")

    from_date = request.args.get("from_date", first_day)
    to_date = request.args.get("to_date", last_day)

    conn = get_conn()
    c = conn.cursor()

    if session.get("role") == "admin":
        c.execute("""
            SELECT id, category, amount, note, date
            FROM expenses
            WHERE date BETWEEN %s AND %s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date))
    else:
        c.execute("""
            SELECT id, category, amount, note, date
            FROM expenses
            WHERE date BETWEEN %s AND %s
            AND company_code=%s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date, session["company_code"]))

    records = c.fetchall()
    conn.close()

    rows = ""
    for r in records:
        rows += f"""
        <tr>
            <td>{str(r[4])[:10]}</td>
            <td>{r[1]}</td>
            <td>RM {float(r[2]):,.2f}</td>
            <td>{r[3]}</td>
            <td>
                <a href="/edit-expense/{r[0]}">Edit</a> |
                <a href="/delete-expense/{r[0]}" onclick="return confirm('Delete this expense?')">Delete</a>
            </td>
        </tr>
        """

    return f"""
    <h1>Expenses List</h1>

    <form method="GET">
        <label>From Date:</label>
        <input type="date" name="from_date" value="{from_date}" required>

        <label>To Date:</label>
        <input type="date" name="to_date" value="{to_date}" required>

        <button type="submit">Search</button>
    </form>

    <br>

    <table border="1" cellpadding="10">
        <tr>
            <th>Date</th>
            <th>Category</th>
            <th>Amount</th>
            <th>Note</th>
            <th>Action</th>
        </tr>
        {rows}
    </table>

    <br>
    <a href="/">Back to Dashboard</a>
    """

@app.route("/edit-expense/<int:expense_id>", methods=["GET", "POST"])
def edit_expense(expense_id):
    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":
        expense_date = request.form["expense_date"]
        category = request.form["category"]
        amount = float(request.form["amount"].replace(",", ""))
        note = request.form["note"]

        c.execute("""
            UPDATE expenses
            SET date=%s, category=%s, amount=%s, note=%s
            WHERE id=%s
        """, (expense_date, category, amount, note, expense_id))

        conn.commit()
        conn.close()
        return redirect("/expenses-list")

    c.execute("SELECT id, category, amount, note, date FROM expenses WHERE id=%s", (expense_id,))
    expense = c.fetchone()
    conn.close()

    return f"""
    <h1>Edit Expense</h1>

    <form method="POST">
        <label>Date:</label><br>
        <input type="date" name="expense_date" value="{str(expense[4])[:10]}" required><br><br>

        <label>Category:</label><br>
        <input type="text" name="category" value="{expense[1]}" required><br><br>

        <label>Amount (RM):</label><br>
        <input type="text" name="amount" value="{float(expense[2]):,.2f}" required><br><br>

        <label>Note:</label><br>
        <input type="text" name="note" value="{expense[3]}"><br><br>

        <button type="submit">Update Expense</button>
    </form>

    <br>
    <a href="/expenses-list">Back to Expenses List</a>
    """

@app.route("/delete-expense/<int:expense_id>")
def delete_expense(expense_id):
    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM expenses WHERE id=%s", (expense_id,))
    conn.commit()
    conn.close()

    return redirect("/expenses-list")


@app.route("/report", methods=["GET", "POST"])
def report():
    if not session.get("logged_in"):
        return redirect("/login")

    sales_total = 0
    expenses_total = 0
    salary_total = 0
    total_expenses = 0
    profit = 0
    from_date = ""
    to_date = ""
    report_generated = False

    if request.method == "POST":
        from_date = request.form["from_date"]
        to_date = request.form["to_date"]
        report_generated = True

        conn = get_conn()
        c = conn.cursor()

        if session.get("role") == "admin":
            c.execute("""
                SELECT COALESCE(SUM(amount),0)
                FROM sales
                WHERE date BETWEEN %s AND %s
            """, (from_date, to_date))
        else:
            c.execute("""
                SELECT COALESCE(SUM(amount),0)
                FROM sales
                WHERE date BETWEEN %s AND %s
                AND company_code=%s
            """, (from_date, to_date, session["company_code"]))

        sales_total = c.fetchone()[0]

        if session.get("role") == "admin":
            c.execute("""
                SELECT COALESCE(SUM(amount),0)
                FROM expenses
                WHERE date BETWEEN %s AND %s
            """, (from_date, to_date))
        else:
            c.execute("""
                SELECT COALESCE(SUM(amount),0)
                FROM expenses
                WHERE date BETWEEN %s AND %s
                AND company_code=%s
            """, (from_date, to_date, session["company_code"]))

        expenses_total = c.fetchone()[0]

        if session.get("role") == "admin":
            c.execute("""
                SELECT COALESCE(SUM(amount),0)
                FROM salaries
                WHERE date BETWEEN %s AND %s
            """, (from_date, to_date))
        else:
            c.execute("""
                SELECT COALESCE(SUM(amount),0)
                FROM salaries
                WHERE date BETWEEN %s AND %s
                AND company_code=%s
            """, (from_date, to_date, session["company_code"]))

        salary_total = c.fetchone()[0]

        conn.close()

        total_expenses = expenses_total + salary_total
        profit = sales_total - total_expenses

    return f"""
    <h1>Profit & Loss Report</h1>

    <form method="POST">
        <label>From Date:</label><br>
        <input type="date" name="from_date" value="{from_date}" required><br><br>

        <label>To Date:</label><br>
        <input type="date" name="to_date" value="{to_date}" required><br><br>

        <button type="submit">Generate Report</button>
    </form>

    <br>

    {f'''
    <h2>Report From {from_date} To {to_date}</h2>

    <table border="1" cellpadding="10">
        <tr>
            <th>Total Sales</th>
            <td>RM {sales_total:,.2f}</td>
        </tr>
        <tr>
            <th>Expenses</th>
            <td>RM {expenses_total:,.2f}</td>
        </tr>
        <tr>
            <th>Staff Salaries</th>
            <td>RM {salary_total:,.2f}</td>
        </tr>
        <tr>
            <th>Total Expenses</th>
            <td>RM {total_expenses:,.2f}</td>
        </tr>
        <tr>
            <th>Profit / Loss</th>
            <td>RM {profit:,.2f}</td>
        </tr>
    </table>
    ''' if report_generated else ''}

    <br>
    <a href="/">Back to Dashboard</a>
    """

@app.route("/reports")
def reports():
    if not session.get("logged_in"):
        return redirect("/login")

    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Reports</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                background: #f4f6f8;
                margin: 0;
                padding: 40px;
            }

            .container {
                max-width: 850px;
                margin: auto;
            }

            .header {
                background: #2f2f2f;
                color: white;
                padding: 30px;
                border-radius: 12px;
                margin-bottom: 30px;
            }

            .header h1 {
                margin: 0;
                font-size: 36px;
            }

            .header p {
                margin-top: 8px;
                color: #ddd;
            }

            .report-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(230px, 1fr));
                gap: 20px;
            }

            .report-card {
                background: white;
                padding: 25px;
                border-radius: 12px;
                box-shadow: 0 4px 12px rgba(0,0,0,0.08);
                text-decoration: none;
                color: #222;
                transition: 0.2s;
            }

            .report-card:hover {
                transform: translateY(-4px);
                box-shadow: 0 6px 18px rgba(0,0,0,0.15);
            }

            .report-card h2 {
                margin: 0 0 10px;
                font-size: 22px;
            }

            .report-card p {
                margin: 0;
                color: #666;
                font-size: 14px;
            }

            .back {
                display: inline-block;
                margin-top: 30px;
                background: #333;
                color: white;
                padding: 12px 18px;
                border-radius: 8px;
                text-decoration: none;
            }

            .back:hover {
                background: #111;
            }
        </style>
    </head>

    <body>
        <div class="container">

            <div class="header">
                <h1>Reports</h1>
                <p>View sales, expenses, salary and financial reports</p>
            </div>

            <div class="report-grid">

                <a href="/sales-list" class="report-card">
                    <h2>Sales List</h2>
                    <p>View sales records by date range</p>
                </a>

                <a href="/expenses-list" class="report-card">
                    <h2>Expenses List</h2>
                    <p>View business expenses by date range</p>
                </a>

                <a href="/salary-list" class="report-card">
                    <h2>Salary List</h2>
                    <p>View staff salary records by date range</p>
                </a>

                <a href="/report" class="report-card">
                    <h2>Monthly P&L</h2>
                    <p>Check profit and loss summary</p>
                </a>

                <a href="/search" class="report-card">
                    <h2>Detailed Report</h2>
                    <p>View sales, expenses and salary details together</p>
                </a>

                <a href="/backup-excel" class="report-card">
                    <h2>Backup Data</h2>
                    <p>Download all sales, expenses and salary data</p>
                </a>

                <form action="/restore-backup"
                      method="POST"
                      enctype="multipart/form-data"
                      class="report-card">

                    <h2>Restore Backup</h2>

                    <p>Upload backup Excel and restore data</p>

                    <input type="file"
                           name="backup_file"
                           accept=".xlsx"
                           required>

                    <br><br>

                    <button type="submit">
                        Restore Data
                    </button>

                </form>

            </div>

            <a href="/" class="back">Back to Dashboard</a>

        </div>
    
    <script>
    document.addEventListener("DOMContentLoaded", function() {

        const urlParams = new URLSearchParams(window.location.search);

        if (urlParams.get("msg") === "restore_success") {

            const toast = document.createElement("div");

            toast.innerHTML = "✅ Backup restored successfully!";

            toast.style.position = "fixed";
            toast.style.top = "50%";
            toast.style.left = "50%";
            toast.style.transform = "translate(-50%, -50%)";

            toast.style.background = "#28a745";
            toast.style.color = "#fff";
            toast.style.padding = "30px 60px";
            toast.style.borderRadius = "15px";
            toast.style.fontWeight = "bold";
            toast.style.fontSize = "24px";
            toast.style.boxShadow = "0 10px 30px rgba(0,0,0,0.4)";
            toast.style.zIndex = "99999";

            document.body.appendChild(toast);

            setTimeout(() => {
                toast.remove();
            }, 3000);

            window.history.replaceState({}, document.title, "/reports");
        }

    });
    </script>

    </body>
    </html>
    """

@app.route("/search", methods=["GET", "POST"])
def search():
    if not session.get("logged_in"):
        return redirect("/login")

    keyword = ""
    sales_results = []
    expenses_results = []
    salary_results = []

    today = datetime.now()
    from_date = today.replace(day=1).strftime("%Y-%m-%d")
    to_date = today.strftime("%Y-%m-%d")

    if request.method == "POST":

        from_date = request.form["from_date"]
        to_date = request.form["to_date"]

    conn = get_conn()
    c = conn.cursor()

    # Sales
    if session.get("role") == "admin":
        c.execute("""
            SELECT date, customer, amount, staff
            FROM sales
            WHERE date BETWEEN %s AND %s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date))
    else:
        c.execute("""
            SELECT date, customer, amount, staff
            FROM sales
            WHERE date BETWEEN %s AND %s
            AND company_code=%s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date, session["company_code"]))

    sales_results = c.fetchall()

    # Expenses
    if session.get("role") == "admin":
        c.execute("""
            SELECT date, category, amount, note
            FROM expenses
            WHERE date BETWEEN %s AND %s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date))
    else:
        c.execute("""
            SELECT date, category, amount, note
            FROM expenses
            WHERE date BETWEEN %s AND %s
            AND company_code=%s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date, session["company_code"]))

    expenses_results = c.fetchall()

    # Salaries
    if session.get("role") == "admin":
        c.execute("""
            SELECT date, staff, amount, month
            FROM salaries
            WHERE date BETWEEN %s AND %s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date))
    else:
        c.execute("""
            SELECT date, staff, amount, month
            FROM salaries
            WHERE date BETWEEN %s AND %s
            AND company_code=%s
            ORDER BY date DESC, id DESC
        """, (from_date, to_date, session["company_code"]))

    salary_results = c.fetchall()

    conn.close()

    sales_rows = ""
    for r in sales_results:
        sales_rows += f"""
        <tr>
            <td>{str(r[0])[:10]}</td>
            <td>{r[1]}</td>
            <td>RM {float(r[2]):,.2f}</td>
            <td>{r[3]}</td>
        </tr>
        """

    expenses_rows = ""
    for r in expenses_results:
        expenses_rows += f"""
        <tr>
            <td>{str(r[0])[:10]}</td>
            <td>{r[1]}</td>
            <td>RM {float(r[2]):,.2f}</td>
            <td>{r[3]}</td>
        </tr>
        """

    salary_rows = ""
    for r in salary_results:
        salary_rows += f"""
        <tr>
            <td>{str(r[0])[:10]}</td>
            <td>{r[1]}</td>
            <td>RM {float(r[2]):,.2f}</td>
            <td>{r[3]}</td>
        </tr>
        """

    return f"""
    <h1>Detailed Report</h1>

    <form method="POST">
        <label>From Date:</label>
        <input type="date" name="from_date" value="{from_date}" required>

        <label>To Date:</label>
        <input type="date" name="to_date" value="{to_date}" required>

        <br><br>

        <button type="submit">Search</button>

        <a href="/export-excel">
            <button type="button">Export Excel</button>
        </a>

        <a href="/export-pdf">
            <button type="button">Export PDF</button>
        </a>


    </form>

    <hr>

    <h2>Sales Results</h2>
    <table border="1" cellpadding="10">
        <tr>
            <th>Date</th>
            <th>Customer</th>
            <th>Amount</th>
            <th>Staff</th>
        </tr>
        {sales_rows}
    </table>

    <h2>Expenses Results</h2>
    <table border="1" cellpadding="10">
        <tr>
            <th>Date</th>
            <th>Category</th>
            <th>Amount</th>
            <th>Note</th>
        </tr>
        {expenses_rows}
    </table>

    <h2>Salary Results</h2>
    <table border="1" cellpadding="10">
        <tr>
            <th>Date</th>
            <th>Staff</th>
            <th>Amount</th>
            <th>Month</th>
        </tr>
        {salary_rows}
    </table>

    <br>
    <a href="/">Back to Dashboard</a>
    """

@app.route("/export-excel")
def export_excel():

    wb = Workbook()

    ws = wb.active
    ws.title = "Detailed Report"

    ws.append(["Date", "Type", "Description", "Amount"])

    conn = get_conn()
    c = conn.cursor()

    # Sales
    c.execute("""
    SELECT date, customer, amount, staff
    FROM sales
    ORDER BY date DESC
    """)
    sales = c.fetchall()

    for r in sales:
        ws.append([str(r[0])[:10], "Sale", f"{r[1]} / Staff: {r[3]}", float(r[2])])

    # Expenses
    c.execute("""
    SELECT date, category, amount, note
    FROM expenses
    ORDER BY date DESC
    """)
    expenses = c.fetchall()

    for r in expenses:
        ws.append([str(r[0])[:10], "Expense", f"{r[1]} / {r[3]}", float(r[2])])

    # Salary
    c.execute("""
    SELECT date, staff, amount, month
    FROM salaries
    ORDER BY date DESC
    """)
    salaries = c.fetchall()

    for r in salaries:
        ws.append([str(r[0])[:10], "Salary", f"{r[1]} / Month: {r[3]}", float(r[2])])

    conn.close()

    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '#,##0.00'

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Detailed_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/backup-excel")
def backup_excel():
    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    wb = Workbook()

    # Sales Sheet
    ws = wb.active
    ws.title = "Sales"
    ws.append(["ID", "Date", "Customer", "Amount", "Staff"])

    c.execute("SELECT id, date, customer, amount, staff FROM sales ORDER BY date DESC")
    for r in c.fetchall():
        ws.append([r[0], str(r[1])[:10], r[2], float(r[3]), r[4]])

    # Expenses Sheet
    ws2 = wb.create_sheet("Expenses")
    ws2.append(["ID", "Date", "Category", "Amount", "Note"])

    c.execute("SELECT id, date, category, amount, note FROM expenses ORDER BY date DESC")
    for r in c.fetchall():
        ws2.append([r[0], str(r[1])[:10], r[2], float(r[3]), r[4]])

    # Salaries Sheet
    ws3 = wb.create_sheet("Salaries")
    ws3.append(["ID", "Date", "Staff", "Amount", "Month"])

    c.execute("SELECT id, date, staff, amount, month FROM salaries ORDER BY date DESC")
    for r in c.fetchall():
        ws3.append([r[0], str(r[1])[:10], r[2], float(r[3]), r[4]])

    conn.close()

    # Format amount columns
    for sheet in [ws, ws2, ws3]:
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = '#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="optic_shop_backup.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/restore-backup", methods=["POST"])
def restore_backup():
    if not session.get("logged_in"):
        return redirect("/login")

    file = request.files.get("backup_file")

    if not file:
        return "No backup file uploaded"

    wb = load_workbook(file)

    conn = get_conn()
    c = conn.cursor()

    try:
        # Clear old data
        c.execute("DELETE FROM sales")
        c.execute("DELETE FROM expenses")
        c.execute("DELETE FROM salaries")

        # Restore Sales
        if "Sales" in wb.sheetnames:
            ws = wb["Sales"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]:
                    continue

                date = row[1]
                customer = row[2]
                amount = row[3]
                staff = row[4]

                c.execute("""
                    INSERT INTO sales (date, customer, amount, staff)
                    VALUES (%s, %s, %s, %s)
                """, (date, customer, amount, staff))

        # Restore Expenses
        if "Expenses" in wb.sheetnames:
            ws = wb["Expenses"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]:
                    continue

                date = row[1]
                category = row[2]
                amount = row[3]
                note = row[4]

                c.execute("""
                    INSERT INTO expenses (date, category, amount, note)
                    VALUES (%s, %s, %s, %s)
                """, (date, category, amount, note))

        # Restore Salaries
        if "Salaries" in wb.sheetnames:
            ws = wb["Salaries"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]:
                    continue

                date = row[1]
                staff = row[2]
                amount = row[3]
                month = row[4]

                c.execute("""
                    INSERT INTO salaries (date, staff, amount, month)
                    VALUES (%s, %s, %s, %s)
                """, (date, staff, amount, month))

        conn.commit()

    except Exception as e:
        conn.rollback()
        conn.close()
        return f"Restore failed: {e}"

    conn.close()

    return redirect("/reports?msg=restore_success")

@app.route("/export-pdf")
def export_pdf():

    conn = get_conn()
    c = conn.cursor()

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()

    elements = []

    title = Paragraph(
        "<b>OPTIC SHOP REPORT</b>",
        styles["Title"]
    )

    elements.append(title)
    elements.append(Spacer(1,20))
    elements.append(
        Paragraph("<b>SALES</b>", styles["Heading2"])
    )

    c.execute("""
        SELECT date, customer, staff, amount
        FROM sales
        ORDER BY date DESC
    """)

    sales_data = [
        ["Date","Customer","Staff","Amount"]
    ]

    for r in c.fetchall():

        sales_data.append([
            str(r[0])[:10],
            r[1],
            r[2],
            f"{float(r[3]):,.2f}"
        ])

    sales_table = Table(
        sales_data,
        colWidths=[90,220,80,100]
    )

    sales_table.setStyle(TableStyle([
        ('GRID',(0,0),(-1,-1),1,colors.black),
        ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('ALIGN',(3,1),(3,-1),'RIGHT')
    ]))

    elements.append(sales_table)
    elements.append(Spacer(1,20))
    elements.append(
        Paragraph("<b>EXPENSES</b>", styles["Heading2"])
    )

    c.execute("""
        SELECT date, category, note, amount
        FROM expenses
        ORDER BY date DESC
    """)

    expense_data = [
        ["Date","Category","Note","Amount"]
    ]

    for r in c.fetchall():

        expense_data.append([
            str(r[0])[:10],
            r[1],
            r[2],
            f"{float(r[3]):,.2f}"
        ])

    expense_table = Table(
        expense_data,
        colWidths=[90,140,180,80]
    )

    expense_table.setStyle(TableStyle([
        ('GRID',(0,0),(-1,-1),1,colors.black),
        ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('ALIGN',(3,1),(3,-1),'RIGHT')
    ]))

    elements.append(expense_table)
    elements.append(Spacer(1,20))
    elements.append(
        Paragraph("<b>SALARIES</b>", styles["Heading2"])
    )

    c.execute("""
        SELECT date, staff, month, amount
        FROM salaries
        ORDER BY date DESC
    """)

    salary_data = [
        ["Date","Staff","Month","Amount"]
    ]

    for r in c.fetchall():

        salary_data.append([
            str(r[0])[:10],
            r[1],
            r[2],
            f"{float(r[3]):,.2f}"
        ])

    salary_table = Table(
        salary_data,
        colWidths=[90,220,100,80]
    )

    salary_table.setStyle(TableStyle([
        ('GRID',(0,0),(-1,-1),1,colors.black),
        ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('ALIGN',(3,1),(3,-1),'RIGHT')
    ]))

    elements.append(salary_table)

    c.execute("SELECT COALESCE(SUM(amount),0) FROM sales")
    total_sales = c.fetchone()[0]

    c.execute("SELECT COALESCE(SUM(amount),0) FROM expenses")
    total_expenses = c.fetchone()[0]

    c.execute("SELECT COALESCE(SUM(amount),0) FROM salaries")
    total_salary = c.fetchone()[0]

    profit = total_sales - total_expenses - total_salary

    elements.append(Spacer(1,25))

    summary = [
        ["Total Sales", f"RM {total_sales:,.2f}"],
        ["Total Expenses", f"RM {total_expenses:,.2f}"],
        ["Total Salary", f"RM {total_salary:,.2f}"],
        ["Net Profit", f"RM {profit:,.2f}"]
    ]

    summary_table = Table(
        summary,
        colWidths=[180,180]
    )

    summary_table.setStyle(TableStyle([
        ('GRID',(0,0),(-1,-1),1,colors.black),
        ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold')
    ]))

    elements.append(summary_table)

    doc.build(elements)

    conn.close()

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="Detailed_Report.pdf",
        mimetype="application/pdf"
    )

@app.route("/permissions")
def permissions():

    if session.get("role") not in ["admin", "owner"]:
        return "Access Denied"

    conn = get_conn()
    c = conn.cursor()

    if session.get("role") == "admin":
        c.execute("""
            SELECT id, username, role
            FROM users
            ORDER BY username
    """)
    else:
        c.execute("""
            SELECT id, username, role
            FROM users
            WHERE role != 'admin'
            AND company_code=%s      
            ORDER BY username
        """, (session["company_code"],))

    users = c.fetchall()
    conn.close()

    rows = ""

    for u in users:
        rows += f"""
        <tr>
            <td>{u[0]}</td>
            <td>{u[1]}</td>
            <td>{u[2]}</td>
            <td>
                <a href="/edit-user/{u[0]}">Edit Permission</a>
                 |
                <a href="/reset-password/{u[0]}">Reset Password</a>
                 |
                <a href="/toggle-user/{u[0]}">Enable / Disable</a>
                 |
                <a href="/delete-user/{u[0]}" onclick="return confirm('Delete this user?')">Delete</a>
            </td>
        </tr>
        """

    return f"""
    <h1>Permission Management</h1>

    <a href="/add-user">
        <button>Add User</button>
    </a>
    <br><br>

    <table border="1" cellpadding="10">
        <tr>
            <th>ID</th>
            <th>Username</th>
            <th>Role</th>
            <th>Action</th>
        </tr>

        {rows}
    </table>

    <br>
    <a href="/">Back Dashboard</a>
    """

@app.route("/companies")
def companies():

    if session.get("role") != "admin":
        return "Access Denied"

    conn = get_conn()
    c = conn.cursor()

    c.execute("""
        SELECT id,
               company_code,
               company_name,
               address,
               phone,
               is_active,
               expiry_date
        FROM companies
        ORDER BY company_name
    """)

    companies_list = c.fetchall()
    conn.close()

    rows = ""

    for company in companies_list:

        status = "ACTIVE" if company[5] else "SUSPENDED"
        expiry = company[6] or "-"

        action_text = "Suspend" if company[5] else "Activate"

        rows += f"""
        <tr>
            <td>{company[0]}</td>
            <td>{company[1]}</td>
            <td>{company[2]}</td>
            <td>{company[3] or ''}</td>
            <td>{company[4] or ''}</td>
            <td>{status}</td>
            <td>{expiry}</td>

            <td>
                <a href="/toggle-company/{company[0]}">{action_text}</a>
                |
                <a href="/edit-company-expiry/{company[0]}">Expiry</a>
                |
                <a href="/reset-company-data?company_code={company[1]}"
                   onclick="return confirm('Reset all data for this company?')">
                   Reset Data
                </a>
            </td>
        </tr>
        """

    return f"""
    <h1>Company Management</h1>

    <a href="/add-company">
        <button>Add Company</button>
    </a>

    <br><br>

    <table border="1" cellpadding="10">
        <tr>
            <th>ID</th>
            <th>Company Code</th>
            <th>Company Name</th>
            <th>Address</th>
            <th>Phone</th>
            <th>Status</th>
            <th>Expiry Date</th>
            <th>Actions</th>
        </tr>

        {rows}

    </table>

    <br>

    <a href="/">Back Dashboard</a>
    """

@app.route("/edit-company-expiry/<int:company_id>", methods=["GET", "POST"])
def edit_company_expiry(company_id):

    if not session.get("logged_in"):
        return redirect("/login")

    if session.get("role") != "admin":
        return "Access denied"

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":
        expiry_date = request.form["expiry_date"] or None

        c.execute("""
            UPDATE companies
            SET expiry_date=%s
            WHERE id=%s
        """, (expiry_date, company_id))

        conn.commit()
        conn.close()

        return redirect("/companies")

    c.execute("""
        SELECT company_name, expiry_date
        FROM companies
        WHERE id=%s
    """, (company_id,))

    company = c.fetchone()
    conn.close()

    return f"""
    <h1>Edit Company Expiry</h1>

    <form method="POST">
        Company:<br>
        <b>{company[0]}</b><br><br>

        Expiry Date:<br>
        <input type="date" name="expiry_date" value="{company[1] or ''}"><br><br>

        <button type="submit">Save Expiry</button>
    </form>

    <br>
    <a href="/companies">Back Companies</a>
    """

@app.route("/add-company", methods=["GET", "POST"])
def add_company():

    if session.get("role") != "admin":
        return "Access Denied"

    if request.method == "POST":
        company_code = request.form["company_code"]
        company_name = request.form["company_name"]
        address = request.form["address"]
        phone = request.form["phone"]

        conn = get_conn()
        c = conn.cursor()

        c.execute("""
            INSERT INTO companies
            (company_code, company_name, address, phone, is_active)
            VALUES (%s, %s, %s, %s, TRUE)
        """, (company_code, company_name, address, phone))

        conn.commit()
        conn.close()

        return redirect("/companies")

    return """
    <h1>Add Company</h1>

    <form method="POST">
        Company Code:<br>
        <input type="text" name="company_code" required><br><br>

        Company Name:<br>
        <input type="text" name="company_name" required><br><br>

        Address:<br>
        <input type="text" name="address"><br><br>

        Phone:<br>
        <input type="text" name="phone"><br><br>

        <button type="submit">Create Company</button>
    </form>

    <br>
    <a href="/companies">Back</a>
    """

@app.route("/add-user", methods=["GET", "POST"])
def add_user():

    if session.get("role") not in ["admin", "owner"]:
        return "Access Denied"

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        role = request.form["role"]

        if session.get("role") == "owner" and role in ["admin", "owner"]:
            return "Access Denied"

        conn = get_conn()
        c = conn.cursor()

        c.execute("""
            INSERT INTO users (username, password, role, company_code, is_active)
            VALUES (%s, %s, %s, %s, TRUE)
        """, (username, password, role, session["company_code"]))

        conn.commit()
        conn.close()

        return redirect("/permissions")

    return """
    <h1>Add User</h1>

    <form method="POST">
        Username:<br>
        <input type="text" name="username" required><br><br>

        Password:<br>
        <input type="text" name="password" required><br><br>

        Role:<br>
        <select name="role">
            <option value="staff">Staff</option>
            <option value="manager">Manager</option>
            <option value="owner">Owner</option>
        </select><br><br>

        <button type="submit">Create User</button>
    </form>

    <br>
    <a href="/permissions">Back</a>
    """

@app.route("/edit-user/<int:user_id>", methods=["GET", "POST"])
def edit_user(user_id):

    if session.get("role") not in ["admin", "owner"]:
        return "Access Denied"

    conn = get_conn()
    c = conn.cursor()

    if session.get("role") == "owner":
        c.execute(
            "SELECT role FROM users WHERE id=%s",
            (user_id,)
        )
        target = c.fetchone()

        if not target or target[0] == "admin":
            conn.close()
            return "Access Denied"

    if request.method == "POST":
        role = request.form["role"]

        permissions = [
            "can_add_sales", "can_edit_sales", "can_delete_sales",
            "can_add_expenses", "can_edit_expenses", "can_delete_expenses",
            "can_add_salary", "can_edit_salary", "can_delete_salary",
            "can_view_reports", "can_export", "can_backup", "can_restore",
            "is_active"
        ]

        values = [role]
        for p in permissions:
            values.append(p in request.form)

        values.append(user_id)

        c.execute("""
            UPDATE users SET
                role=%s,
                can_add_sales=%s,
                can_edit_sales=%s,
                can_delete_sales=%s,
                can_add_expenses=%s,
                can_edit_expenses=%s,
                can_delete_expenses=%s,
                can_add_salary=%s,
                can_edit_salary=%s,
                can_delete_salary=%s,
                can_view_reports=%s,
                can_export=%s,
                can_backup=%s,
                can_restore=%s,
                is_active=%s
            WHERE id=%s
        """, values)

        conn.commit()
        conn.close()

        return redirect("/permissions")

    c.execute("""
        SELECT id, username, role,
               can_add_sales, can_edit_sales, can_delete_sales,
               can_add_expenses, can_edit_expenses, can_delete_expenses,
               can_add_salary, can_edit_salary, can_delete_salary,
               can_view_reports, can_export, can_backup, can_restore,
               is_active
        FROM users
        WHERE id=%s
    """, (user_id,))

    u = c.fetchone()
    conn.close()

    def checked(value):
        return "checked" if value else ""

    return f"""
    <h1>Edit Permission: {u[1]}</h1>

    <form method="POST">
        <label>Role:</label>
        <select name="role">
            <option value="admin" {"selected" if u[2]=="admin" else ""}>Admin</option>
            <option value="manager" {"selected" if u[2]=="manager" else ""}>Manager</option>
            <option value="staff" {"selected" if u[2]=="staff" else ""}>Staff</option>
        </select>

        <h3>Sales</h3>
        <label><input type="checkbox" name="can_add_sales" {checked(u[3])}> Add Sales</label><br>
        <label><input type="checkbox" name="can_edit_sales" {checked(u[4])}> Edit Sales</label><br>
        <label><input type="checkbox" name="can_delete_sales" {checked(u[5])}> Delete Sales</label><br>

        <h3>Expenses</h3>
        <label><input type="checkbox" name="can_add_expenses" {checked(u[6])}> Add Expenses</label><br>
        <label><input type="checkbox" name="can_edit_expenses" {checked(u[7])}> Edit Expenses</label><br>
        <label><input type="checkbox" name="can_delete_expenses" {checked(u[8])}> Delete Expenses</label><br>

        <h3>Salary</h3>
        <label><input type="checkbox" name="can_add_salary" {checked(u[9])}> Add Salary</label><br>
        <label><input type="checkbox" name="can_edit_salary" {checked(u[10])}> Edit Salary</label><br>
        <label><input type="checkbox" name="can_delete_salary" {checked(u[11])}> Delete Salary</label><br>

        <h3>Reports</h3>
        <label><input type="checkbox" name="can_view_reports" {checked(u[12])}> View Reports</label><br>
        <label><input type="checkbox" name="can_export" {checked(u[13])}> Export</label><br>
        <label><input type="checkbox" name="can_backup" {checked(u[14])}> Backup</label><br>
        <label><input type="checkbox" name="can_restore" {checked(u[15])}> Restore</label><br>

        <h3>Status</h3>
        <label><input type="checkbox" name="is_active" {checked(u[16])}> Active</label><br><br>

        <button type="submit">Save Permission</button>
    </form>

    <br>
    <a href="/permissions">Back</a>
    """

@app.route("/delete-user/<int:user_id>")
def delete_user(user_id):

    if session.get("role") not in ["admin", "owner"]:
        return "Access Denied"

    conn = get_conn()
    c = conn.cursor()

    if session.get("role") == "owner":
        c.execute(
            "SELECT role FROM users WHERE id=%s",
            (user_id,)
        )
        target = c.fetchone()

        if target and target[0] == "admin":
            conn.close()
            return "Access Denied"

    c.execute(
        "DELETE FROM users WHERE id=%s",
        (user_id,)
    )

    conn.commit()
    conn.close()

    return redirect("/permissions")

@app.route("/reset-password/<int:user_id>", methods=["GET", "POST"])
def reset_password(user_id):

    if session.get("role") not in ["admin", "owner"]:
        return "Access Denied"

    conn = get_conn()
    c = conn.cursor()

    if session.get("role") == "owner":
        c.execute("SELECT role FROM users WHERE id=%s", (user_id,))
        target = c.fetchone()

        if target and target[0] == "admin":
            conn.close()
            return "Access Denied"

    if request.method == "POST":
        new_password = request.form["password"]

        c.execute(
            "UPDATE users SET password=%s WHERE id=%s",
            (new_password, user_id)
        )

        conn.commit()
        conn.close()

        return redirect("/permissions")

    conn.close()

    return """
    <h1>Reset Password</h1>

    <form method="POST">
        New Password:<br>
        <input type="text" name="password" required><br><br>

        <button type="submit">Save Password</button>
    </form>

    <br>
    <a href="/permissions">Back</a>
    """

@app.route("/toggle-user/<int:user_id>")
def toggle_user(user_id):

    if session.get("role") not in ["admin", "owner"]:
        return "Access Denied"

    conn = get_conn()
    c = conn.cursor()

    # owner不能动admin
    if session.get("role") == "owner":
        c.execute(
            "SELECT role FROM users WHERE id=%s",
            (user_id,)
        )
        target = c.fetchone()

        if target and target[0] == "admin":
            conn.close()
            return "Access Denied"

    c.execute(
        "SELECT is_active FROM users WHERE id=%s",
        (user_id,)
    )

    user = c.fetchone()

    if user:
        new_status = not user[0]

        c.execute(
            "UPDATE users SET is_active=%s WHERE id=%s",
            (new_status, user_id)
        )

        conn.commit()

    conn.close()

    return redirect("/permissions")

@app.route("/admin")
def admin_dashboard():
    if not session.get("logged_in"):
        return redirect("/login")

    if session.get("role") != "admin":
        return "Access Denied"

    return """
    <h1>Super Admin Dashboard</h1>

    <ul>
        <li><a href="/register-owner">Create Company</a></li>
        <li><a href="/companies">Companies</a></li>
        <li><a href="/support-tickets">Support Tickets</a></li>
        <li><a href="/error-logs">Error Logs</a></li>
        <li><a href="/reset-company-data"><span style="color:red;">Reset Company Data</span></a></li>
        <li><a href="/logout">Logout</a></li>
    </ul>
    """

@app.route("/support", methods=["GET", "POST"])
def support():

    if not session.get("logged_in"):
        return redirect("/login")

    if request.method == "POST":

        subject = request.form["subject"]
        description = request.form["description"]

        conn = get_conn()
        c = conn.cursor()

        c.execute("""
        INSERT INTO support_tickets
        (company_code, username, subject, description)
        VALUES (%s, %s, %s, %s)
        """, (
            session["company_code"],
            session["username"],
            subject,
            description
        ))

        conn.commit()
        conn.close()

        return """
        <h2>Ticket Submitted</h2>
        <a href="/">Back</a>
        """

    return """
    <h1>Report Problem</h1>

    <form method="POST">

        Subject<br>
        <input type="text" name="subject" required><br><br>

        Description<br>
        <textarea name="description"
                  rows="10"
                  cols="60"
                  required></textarea><br><br>

        <button type="submit">Submit Ticket</button>

    </form>
    """

@app.route("/support-tickets")
def support_tickets():

    if not session.get("logged_in"):
        return redirect("/login")

    if session.get("role") != "admin":
        return "Access Denied"

    conn = get_conn()
    c = conn.cursor()

    c.execute("""
    SELECT
        id,
        company_code,
        username,
        subject,
        status,
        created_at
    FROM support_tickets
    ORDER BY id DESC
    """)

    tickets = c.fetchall()

    conn.close()

    rows = ""

    for t in tickets:

        rows += f"""
        <tr>
            <td>{t[0]}</td>
            <td>{t[1]}</td>
            <td>{t[2]}</td>
            <td>{t[3]}</td>
            <td>{t[4]}</td>
            <td>{t[5]}</td>
        </tr>
        """

    return f"""
    <h1>Support Tickets</h1>

    <table border="1">
        <tr>
            <th>ID</th>
            <th>Company</th>
            <th>User</th>
            <th>Subject</th>
            <th>Status</th>
            <th>Date</th>
        </tr>

        {rows}
    </table>

    <br>
    <a href="/admin">Back</a>
    """

@app.route("/stock", methods=["GET", "POST"])
def stock():

    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":
        item_code = request.form["item_code"]
        item_name = request.form["item_name"]
        cost = float(request.form["cost"] or 0)
        commission = float(request.form["commission"] or 0)
        selling_price = float(request.form["selling_price"] or 0)
        minimum_selling_price = float(request.form["minimum_selling_price"] or 0)
        supplier = request.form["supplier"]

        c.execute("""
            INSERT INTO stock (
                company_code,
                item_code,
                item_name,
                cost,
                commission,
                selling_price,
                minimum_selling_price,
                supplier,
                qty
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            session["company_code"],
            item_code,
            item_name,
            cost,
            commission,
            selling_price,
            minimum_selling_price,
            supplier,
            0
        ))

        conn.commit()

    if session.get("role") == "admin":
        c.execute("""
            SELECT id, item_code, item_name, selling_price, cost, commission,
                   minimum_selling_price, supplier, COALESCE(is_active, TRUE)
            FROM stock
            ORDER BY item_name
        """)
    else:
        c.execute("""
            SELECT id, item_code, item_name, selling_price, cost, commission,
                   minimum_selling_price, supplier, COALESCE(is_active, TRUE)
            FROM stock
            WHERE company_code=%s
            ORDER BY item_name
        """, (session["company_code"],))

    items = c.fetchall()
    conn.close()

    rows = ""
    for i in items:
        status_text = "Active" if i[8] else "Inactive"

        rows += f"""
        <tr>
            <td><a href="/stock-item/{i[0]}">{i[1]}</a></td>
            <td>{i[2]}</td>
            <td align="right">RM {float(i[3] or 0):,.2f}</td>
            <td align="right">RM {float(i[4] or 0):,.2f}</td>
            <td align="right">RM {float(i[5] or 0):,.2f}</td>
            <td align="right">RM {float(i[6] or 0):,.2f}</td>
            <td>{i[7] or '-'}</td>
            <td>{status_text}</td>
        </tr>
        """

    return f"""
    <h1>Stock Item</h1>

    <form method="POST">
        Item Code:<br>
        <input type="text" name="item_code" required><br><br>

        Item Name:<br>
        <input type="text" name="item_name" required><br><br>

        Cost:<br>
        <input type="number" step="0.01" name="cost" value="0"><br><br>

        Commission:<br>
        <input type="number" step="0.01" name="commission" value="0"><br><br>

        Selling Price:<br>
        <input type="number"
               step="0.01"
               name="selling_price"
               value="0"><br><br>

        Minimum Selling Price:<br>
        <input type="number" step="0.01" name="minimum_selling_price" value="0"><br><br>

        Supplier:<br>
        <input type="text" name="supplier"><br><br>

        <button type="submit">Add Stock</button>
    </form>

    <hr>

    <h2>Stock Item</h2>

    <table border="1" cellpadding="8">
        <tr>
            <th>Code</th>
            <th>Item</th>
            <th>Selling Price</th>
            <th>Cost</th>
            <th>Commission</th>
            <th>Min Selling Price</th>
            <th>Supplier</th>
            <th>Status</th>
        </tr>
        {rows}
    </table>

    <br>
    <a href="/">Back Dashboard</a>
    """

@app.route("/stock-adjustment", methods=["GET", "POST"])
def stock_adjustment():

    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":

        item_id = int(request.form["item_id"])
        qty_change = int(request.form["qty_change"])
        adjustment_type = request.form["adjustment_type"]
        reason = request.form["reason"]

        if adjustment_type == "DECREASE":
            qty_change = qty_change * -1

        c.execute("""
            UPDATE stock
            SET qty = qty + %s
            WHERE id=%s
            AND company_code=%s
        """, (
            qty_change,
            item_id,
            session["company_code"]
        ))    

        c.execute("""
            INSERT INTO stock_adjustments (
                company_code,
                adjustment_date,
                item_id,
                qty_change,
                reason,
                created_by
            )
            VALUES (%s,CURRENT_DATE,%s,%s,%s,%s)
        """, (
            session["company_code"],
            item_id,
            qty_change,
            reason,
            session["username"]
        ))

        conn.commit()

    c.execute("""
        SELECT id,item_name
        FROM stock
        WHERE company_code=%s
        AND COALESCE(is_active,TRUE)=TRUE
        ORDER BY item_name
    """, (session["company_code"],))

    items = c.fetchall()

    options = ""

    for i in items:
        options += f"""
        <option value="{i[0]}">
            {i[1]}
        </option>
        """

    c.execute("""
        SELECT
            a.adjustment_date,
            s.item_name,
            a.qty_change,
            a.reason,
            a.created_by
        FROM stock_adjustments a
        JOIN stock s ON s.id=a.item_id
        WHERE a.company_code=%s
        ORDER BY a.id DESC
        LIMIT 100
    """, (session["company_code"],))

    rows = ""

    for r in c.fetchall():

        rows += f"""
        <tr>
            <td>{r[0]}</td>
            <td>{r[1]}</td>
            <td>{r[2]}</td>
            <td>{r[3]}</td>
            <td>{r[4]}</td>
        </tr>
        """

    conn.close()

    return f"""
    <h1>Stock Adjustment</h1>

    <form method="POST">

        Item:<br>
        <select name="item_id">
            {options}
        </select><br><br>

        Adjustment Type:<br>

        <select name="adjustment_type">
            <option value="INCREASE">Increase</option>
            <option value="DECREASE">Decrease</option>
        </select><br><br>

        Qty:<br>

        <input type="number"
               name="qty_change"
               required><br><br>

        Reason:<br>
        <textarea name="reason"
                  required></textarea><br><br>

        <button type="submit">
            Save Adjustment
        </button>

    </form>

    <hr>

    <h2>Audit Trail</h2>

    <table border="1">

        <tr>
            <th>Date</th>
            <th>Item</th>
            <th>Qty Change</th>
            <th>Reason</th>
            <th>User</th>
        </tr>

        {rows}

    </table>

    <br><br>

    <a href="/">
        <button type="button">Back Dashboard</button>
    </a>

    """

@app.route("/stock-balance")
def stock_balance():

    if not session.get("logged_in"):
        return redirect("/login")

    search = request.args.get("search", "")
    supplier_filter = request.args.get("supplier", "")
    status_filter = request.args.get("status", "")

    conn = get_conn()
    c = conn.cursor()

    query = """
        SELECT
            item_code,
            item_name,
            supplier,
            qty,
            COALESCE(is_active, TRUE)
        FROM stock
        WHERE company_code=%s
    """

    params = [session["company_code"]]

    if search:
        query += """
        AND (
            item_code ILIKE %s
            OR item_name ILIKE %s
        )
        """
        params.append(f"%{search}%")
        params.append(f"%{search}%")

    if supplier_filter:
        query += " AND supplier ILIKE %s"
        params.append(f"%{supplier_filter}%")

    if status_filter == "ACTIVE":
        query += " AND COALESCE(is_active, TRUE)=TRUE"

    if status_filter == "INACTIVE":
        query += " AND COALESCE(is_active, TRUE)=FALSE"

    query += " ORDER BY item_name"

    c.execute(query, tuple(params))
    items = c.fetchall()

    c.execute("""
        SELECT DISTINCT supplier
        FROM stock
        WHERE company_code=%s
        AND supplier IS NOT NULL
        AND supplier <> ''
        ORDER BY supplier
    """, (session["company_code"],))

    suppliers = c.fetchall()

    conn.close()

    supplier_options = '<option value="">All Suppliers</option>'

    for s in suppliers:
        selected = "selected" if s[0] == supplier_filter else ""
        supplier_options += f"""
        <option value="{s[0]}" {selected}>{s[0]}</option>
        """

    active_selected = "selected" if status_filter == "ACTIVE" else ""
    inactive_selected = "selected" if status_filter == "INACTIVE" else ""

    rows = ""

    for i in items:
        status_text = "Active" if i[4] else "Inactive"

        low_stock_style = ""
        if i[3] is not None and i[3] <= 0:
            low_stock_style = 'style="color:red;font-weight:bold;"'

        rows += f"""
        <tr>
            <td>{i[0]}</td>
            <td>{i[1]}</td>
            <td>{i[2] or '-'}</td>
            <td align="center" {low_stock_style}>{i[3] or 0}</td>
            <td>{status_text}</td>
        </tr>
        """

    return f"""
    <h1>Check Stock Balance</h1>

    <form method="GET">

        Search Item / Code:
        <input type="text" name="search" value="{search}">

        Supplier:
        <select name="supplier">
            {supplier_options}
        </select>

        Status:
        <select name="status">
            <option value="">All</option>
            <option value="ACTIVE" {active_selected}>Active</option>
            <option value="INACTIVE" {inactive_selected}>Inactive</option>
        </select>

        <button type="submit">Search</button>

    </form>

    <br>

    <table border="1" cellpadding="8">
        <tr>
            <th>Item Code</th>
            <th>Item Name</th>
            <th>Supplier</th>
            <th>Stock Balance</th>
            <th>Status</th>
        </tr>

        {rows}

    </table>

    <br>

    <a href="/">
        <button>Back Dashboard</button>
    </a>
    """

@app.route("/stock-item/<int:item_id>", methods=["GET", "POST"])
def stock_item(item_id):

    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":
        item_name = request.form["item_name"]
        supplier = request.form["supplier"]
        selling_price = float(request.form["selling_price"] or 0)
        minimum_selling_price = float(request.form["minimum_selling_price"] or 0)
        commission = float(request.form["commission"] or 0)
        cost = float(request.form["cost"] or 0)
        is_active = request.form.get("is_active") == "Y"

        c.execute("""
            UPDATE stock
            SET item_name=%s,
                supplier=%s,
                selling_price=%s,
                minimum_selling_price=%s,
                commission=%s,
                cost=%s,
                is_active=%s
            WHERE id=%s
            AND company_code=%s
        """, (
            item_name,
            supplier,
            selling_price,
            minimum_selling_price,
            commission,
            cost,
            is_active,
            item_id,
            session["company_code"]
        ))

        conn.commit()

    c.execute("""
        SELECT id, item_code, item_name, supplier,
               selling_price, minimum_selling_price,
               commission, cost, COALESCE(is_active, TRUE)
        FROM stock
        WHERE id=%s
        AND company_code=%s
    """, (item_id, session["company_code"]))

    item = c.fetchone()
    conn.close()

    if not item:
        return "Item not found"

    active_checked = "checked" if item[8] else ""

    return f"""
    <h1>Stock Item Profile</h1>

    <form method="POST">

        Item Code:<br>
        <input type="text" value="{item[1]}" readonly><br><br>

        Item Name:<br>
        <input type="text" name="item_name" value="{item[2] or ''}" required><br><br>

        Supplier:<br>
        <input type="text" name="supplier" value="{item[3] or ''}"><br><br>

        Selling Price:<br>
        <input type="number" step="0.01" name="selling_price" value="{item[4] or 0}"><br><br>

        Minimum Selling Price:<br>
        <input type="number" step="0.01" name="minimum_selling_price" value="{item[5] or 0}"><br><br>

        Commission:<br>
        <input type="number" step="0.01" name="commission" value="{item[6] or 0}"><br><br>

        Cost:<br>
        <input type="number" step="0.01" name="cost" value="{item[7] or 0}"><br><br>

        <label>
            <input type="checkbox" name="is_active" value="Y" {active_checked}>
            Active
        </label><br><br>

        <button type="submit">Save Item</button>
    </form>

    <br>
    <a href="/stock">Back to Stock Item</a>
    """

@app.route("/suppliers", methods=["GET", "POST"])
def suppliers():

    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":
        supplier_code = request.form["supplier_code"]
        supplier_name = request.form["supplier_name"]
        phone = request.form["phone"]
        address = request.form["address"]
        account_no = request.form["account_no"]

        c.execute("""
            INSERT INTO suppliers (
                company_code,
                supplier_code,
                supplier_name,
                phone,
                address,
                account_no
            )
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            session["company_code"],
            supplier_code,
            supplier_name,
            phone,
            address,
            account_no
        ))

        conn.commit()

    c.execute("""
        SELECT id, supplier_code, supplier_name, phone, address, account_no
        FROM suppliers
        WHERE company_code=%s
        ORDER BY supplier_name
    """, (session["company_code"],))

    suppliers_list = c.fetchall()
    conn.close()

    rows = ""
    for s in suppliers_list:
        rows += f"""
        <tr>
            <td>
                <a href="/supplier-profile/{s[0]}">
                    {s[1]}
                </a>
            </td>
            <td>{s[2]}</td>
            <td>{s[3] or ''}</td>
            <td>{s[4] or ''}</td>
            <td>{s[5] or ''}</td>
        </tr>
        """

    return f"""
    <h1>Supplier Management</h1>

    <form method="POST">
        Supplier Code:<br>
        <input type="text" name="supplier_code" required><br><br>

        Supplier Company Name:<br>
        <input type="text" name="supplier_name" required><br><br>

        Tel No:<br>
        <input type="text" name="phone"><br><br>

        Address:<br>
        <textarea name="address" rows="3" cols="50"></textarea><br><br>

        Account No:<br>
        <input type="text" name="account_no"><br><br>

        <button type="submit">Add Supplier</button>
    </form>

    <hr>

    <h2>Supplier List</h2>

    <table border="1" cellpadding="8">
        <tr>
            <th>Code</th>
            <th>Company Name</th>
            <th>Tel No</th>
            <th>Address</th>
            <th>Account No</th>
        </tr>
        {rows}
    </table>

    <br>
    <a href="/">Back Dashboard</a>
    """

@app.route("/supplier-profile/<int:supplier_id>", methods=["GET", "POST"])
def supplier_profile(supplier_id):

    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":

        supplier_name = request.form["supplier_name"]
        supplier_code = request.form["supplier_code"]
        tel_no = request.form["tel_no"]
        address = request.form["address"]
        account_no = request.form["account_no"]

        is_active = request.form.get("is_active") == "Y"

        c.execute("""
            UPDATE suppliers
            SET supplier_name=%s,
                supplier_code=%s,
                tel_no=%s,
                address=%s,
                account_no=%s,
                is_active=%s
            WHERE id=%s
        """, (
            supplier_name,
            supplier_code,
            tel_no,
            address,
            account_no,
            is_active,
            supplier_id
        ))

        conn.commit()

    c.execute("""
        SELECT
            id,
            supplier_name,
            supplier_code,
            tel_no,
            address,
            account_no,
            COALESCE(is_active, TRUE)
        FROM suppliers
        WHERE id=%s
    """, (supplier_id,))

    supplier = c.fetchone()

    if not supplier:
        conn.close()
        return "Supplier not found"

    active_checked = "checked" if supplier[6] else ""

    c.execute("""
        SELECT
            movement_date,
            reference_no,
            movement_type,
            qty
        FROM stock_movements
        WHERE supplier_id=%s
        ORDER BY movement_date DESC
    """, (supplier_id,))

    purchases = c.fetchall()

    rows = ""

    for p in purchases:

        rows += f"""
        <tr>
            <td>{p[0]}</td>
            <td>{p[1] or ''}</td>
            <td>{p[2]}</td>
            <td>{p[3]}</td>
        </tr>
        """

    conn.close()

    return f"""
    <h1>Supplier Profile</h1>

    <form method="POST">

        Supplier Name:<br>
        <input type="text"
               name="supplier_name"
               value="{supplier[1] or ''}"
               required><br><br>

        Supplier Code:<br>
        <input type="text"
               name="supplier_code"
               value="{supplier[2] or ''}"><br><br>

        Tel No:<br>
        <input type="text"
               name="tel_no"
               value="{supplier[3] or ''}"><br><br>

        Address:<br>
        <textarea name="address">{supplier[4] or ''}</textarea><br><br>

        Account No:<br>
        <input type="text"
               name="account_no"
               value="{supplier[5] or ''}"><br><br>

        <label>
            <input type="checkbox"
                   name="is_active"
                   value="Y"
                   {active_checked}>
            Active Supplier
        </label>

        <br><br>

        <button type="submit">
            Save Supplier
        </button>

    </form>

    <hr>

    <h2>Purchase History</h2>

    <table border="1" cellpadding="8">

        <tr>
            <th>Date</th>
            <th>Reference No</th>
            <th>Type</th>
            <th>Qty</th>
        </tr>

        {rows}

    </table>

    <br>

    <a href="/suppliers">
        Back To Supplier List
    </a>
    """

@app.route("/stock-movement", methods=["GET", "POST"])
def stock_movement():

    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    if request.method == "POST":

        movement_date = request.form["movement_date"]
        movement_type = request.form["movement_type"]
        reference_no = request.form["reference_no"]
        supplier_id = request.form["supplier_id"] or None
        item_id = request.form["item_id"]
        qty = int(request.form["qty"])
        note = request.form["note"]

        c.execute("""
            SELECT qty
            FROM stock
            WHERE id=%s AND company_code=%s
        """, (item_id, session["company_code"]))

        stock_item = c.fetchone()

        if not stock_item:
            conn.close()
            return "Stock item not found"

        current_qty = stock_item[0]

        if movement_type == "OUT" and current_qty < qty:
            conn.close()
            return f"Not enough stock. Current Qty: {current_qty}"

        c.execute("""
            INSERT INTO stock_movements (
                company_code,
                movement_date,
                movement_type,
                reference_no,
                supplier_id,
                item_id,
                qty,
                note,
                status
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            session["company_code"],
            movement_date,
            movement_type,
            reference_no,
            supplier_id,
            item_id,
            qty,
            note,
            "ACTIVE"
        ))

        if movement_type == "IN":
            c.execute("""
                UPDATE stock
                SET qty = qty + %s
                WHERE id=%s AND company_code=%s
            """, (qty, item_id, session["company_code"]))

        if movement_type == "OUT":
            c.execute("""
                UPDATE stock
                SET qty = qty - %s
                WHERE id=%s AND company_code=%s
            """, (qty, item_id, session["company_code"]))

        conn.commit()
        conn.close()

        return redirect("/stock-movement")

    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    ref_search = request.args.get("reference_no", "")
    supplier_filter = request.args.get("supplier_id", "")

    c.execute("""
        SELECT id, supplier_name
        FROM suppliers
        WHERE company_code=%s
        ORDER BY supplier_name
    """, (session["company_code"],))

    suppliers = c.fetchall()

    c.execute("""
        SELECT id, item_name, qty
        FROM stock
        WHERE company_code=%s
        AND COALESCE(is_active,TRUE)=TRUE
        ORDER BY item_name
    """, (session["company_code"],))

    items = c.fetchall()

    supplier_options = '<option value="">All Suppliers</option>'

    for s in suppliers:
        selected = "selected" if str(s[0]) == supplier_filter else ""
        supplier_options += f"""
        <option value="{s[0]}" {selected}>{s[1]}</option>
        """

    item_options = ""

    for i in items:
        item_options += f"""
        <option value="{i[0]}">{i[1]} (Current Qty: {i[2]})</option>
        """

    query = """
        SELECT
            sm.id,
            sm.movement_date,
            sm.movement_type,
            sm.reference_no,
            sp.supplier_name,
            st.item_name,
            sm.qty,
            sm.note,
            sm.status
        FROM stock_movements sm
        LEFT JOIN suppliers sp ON sm.supplier_id = sp.id
        LEFT JOIN stock st ON sm.item_id = st.id
        WHERE sm.company_code=%s
    """

    params = [session["company_code"]]

    if from_date:
        query += " AND sm.movement_date >= %s"
        params.append(from_date)

    if to_date:
        query += " AND sm.movement_date <= %s"
        params.append(to_date)

    if ref_search:
        query += " AND sm.reference_no ILIKE %s"
        params.append(f"%{ref_search}%")

    if supplier_filter:
        query += " AND sm.supplier_id=%s"
        params.append(supplier_filter)

    query += """
    ORDER BY sm.movement_date DESC, sm.id DESC
    """

    c.execute(query, tuple(params))

    records = c.fetchall()

    rows = ""

    for r in records:

        rows += f"""
        <tr>
            <td>{r[1]}</td>
            <td>{r[2]}</td>
            <td>{r[3] or '-'}</td>
            <td>{r[4] or '-'}</td>
            <td>{r[5] or '-'}</td>
            <td align="center">{r[6]}</td>
            <td>{r[7] or ''}</td>
            <td>{r[8] or 'ACTIVE'}</td>
            <td>
        """

        if (r[8] or "ACTIVE") == "ACTIVE":
            rows += f"""
            <a href="/void-stock-movement/{r[0]}"
               onclick="return confirm('Void this movement? Stock quantity will be reversed.')">
               Void
            </a>
            """
        else:
            rows += "VOIDED"

        rows += """
            </td>
        </tr>
        """

    conn.close()

    return f"""
    <h1>Stock Purchase In / Out</h1>

    <form method="POST">

    Date:<br>
    <input type="date" name="movement_date" required><br><br>

    Type:<br>
    <select name="movement_type">
        <option value="IN">Receive Stock</option>
        <option value="OUT">Stock Out</option>
    </select><br><br>

    Reference No:<br>
    <input type="text" name="reference_no"><br><br>

    Supplier:<br>
    <select name="supplier_id">
        <option value="">-</option>
        {supplier_options}
    </select><br><br>

    Item:<br>
    <select name="item_id" required>
        {item_options}
    </select><br><br>

    Qty:<br>
    <input type="number" name="qty" min="1" required><br><br>

    Note:<br>
    <textarea name="note"></textarea><br><br>

    <button type="submit">Save Movement</button>

    </form>

    <hr>

    <form method="GET">

    From:
    <input type="date" name="from_date" value="{from_date}">

    To:
    <input type="date" name="to_date" value="{to_date}">

    Ref No:
    <input type="text" name="reference_no" value="{ref_search}">

    Supplier:
    <select name="supplier_id">
        {supplier_options}
    </select>

    <button type="submit">Search</button>

    </form>

    <br>

    <table border="1" cellpadding="8">

        <tr>
            <th>Date</th>
            <th>Type</th>
            <th>Reference No</th>
            <th>Supplier</th>
            <th>Item</th>
            <th>Qty</th>
            <th>Note</th>
            <th>Status</th>
            <th>Action</th>
        </tr>

        {rows}

    </table>

    <br>
    <a href="/">Back to Dashboard</a>
    """

@app.route("/void-stock-movement/<int:id>")
def void_stock_movement(id):

    if not session.get("logged_in"):
        return redirect("/login")

    conn = get_conn()
    c = conn.cursor()

    c.execute("""
        SELECT movement_type, item_id, qty, status
        FROM stock_movements
        WHERE id=%s AND company_code=%s
    """, (id, session["company_code"]))

    rec = c.fetchone()

    if not rec:
        conn.close()
        return "Stock movement not found"

    movement_type, item_id, qty, status = rec

    if status == "VOID":
        conn.close()
        return redirect("/stock-movement")

    c.execute("""
        SELECT qty
        FROM stock
        WHERE id=%s AND company_code=%s
    """, (item_id, session["company_code"]))

    stock_item = c.fetchone()

    if not stock_item:
        conn.close()
        return "Stock item not found"

    current_qty = stock_item[0] or 0

    if movement_type == "IN":
        new_qty = current_qty - qty
    else:
        new_qty = current_qty + qty

    c.execute("""
        UPDATE stock
        SET qty=%s
        WHERE id=%s AND company_code=%s
    """, (new_qty, item_id, session["company_code"]))

    c.execute("""
        UPDATE stock_movements
        SET status='VOID'
        WHERE id=%s AND company_code=%s
    """, (id, session["company_code"]))

    conn.commit()
    conn.close()

    return redirect("/stock-movement")

@app.route("/reset-company-data", methods=["GET", "POST"])
def reset_company_data():

    if not session.get("logged_in"):
        return redirect("/login")

    if session.get("role") != "admin":
        return "Access denied"

    conn = get_conn()
    c = conn.cursor()

    c.execute("""
        SELECT company_code, company_name
        FROM companies
        ORDER BY company_name
    """)

    companies = c.fetchall()

    company_options = ""

    for company in companies:
        company_options += f"""
        <option value="{company[0]}">
            {company[1]} ({company[0]})
        </option>
        """

    if request.method == "POST":
        company_code = request.form["company_code"]
        confirm = request.form["confirm"]

        if confirm != "RESET":
            return "Please type RESET to confirm"

        conn = get_conn()
        c = conn.cursor()

        tables = [
            "sales",
            "expenses",
            "salaries",
            "stock_movements",
            "stock_adjustments",
            "stock",
            "suppliers"
        ]

        for table in tables:
            c.execute(
                f"DELETE FROM {table} WHERE company_code=%s",
                (company_code,)
            )

        conn.commit()
        conn.close()

        return f"""
        <h1>Reset Completed</h1>

        <p>
        Company <b>{company_code}</b>
        data reset successfully.
        </p>

        <br>

        <a href="/admin">
        Back Super Admin Dashboard
        </a>
        """

    return f"""
    <h1>Reset Company Data</h1>

    <form method="POST">

        Company:<br>
        <select name="company_code" required>{company_options}</select><br><br>

        Type RESET to confirm:<br>
        <input type="text" name="confirm" required><br><br>

        <button type="submit" style="color:red;"
            onclick="return confirm('Are you sure? This will delete sales, stock, suppliers, expenses and salary data.')">
            RESET DATA
        </button>

    </form>

    <br>
    <a href="/admin">Back Super Admin Dashboard</a>
    """

@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect("/login")

if __name__ == "__main__":
    init_db()
    app.run(debug=True)